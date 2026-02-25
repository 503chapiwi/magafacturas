import pdfplumber
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import streamlit as st
import re
import io

# --- HELPER FUNCTIONS ---
def normalize_text(text):
    if not text: return ""
    nfd = unicodedata.normalize('NFD', str(text))
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn').lower()

def safe_float(val):
    """Safely converts any Excel cell value (even with 'Q', '-', or spaces) into a usable number."""
    if val is None: return 0.0
    s = str(val).strip()
    if not s or s == '-': return 0.0
    s = s.replace(',', '') # Handle thousands separators
    s = re.sub(r'[^\d\.\-]', '', s) # Strip everything but numbers and decimals
    try: return float(s)
    except ValueError: return 0.0

def clean_currency(value):
    """Specific parser for OCR-extracted PDF currency amounts."""
    if not value: return 0.0
    raw = str(value).strip()
    raw = re.sub(r'[^\d\.,]', '', raw)
    if not raw: return 0.0
    
    if ',' in raw and '.' in raw:
        raw = raw.replace(',', '') 
    elif ',' in raw and re.search(r',\d{2}$', raw):
        raw = raw.replace(',', '.') 
    else:
        raw = raw.replace(',', '')
        
    try: return float(raw)
    except ValueError: return 0.0

# --- TRUCO CSS PARA TRADUCIR LA INTERFAZ A ESPAÑOL ---
st.markdown("""
    <style>
        .stFileUploader > div > div > div > div > span:first-child {
            display: none;
        }
        .stFileUploader > div > div > div > div::before {
            content: "Arrastre y suelte los archivos aquí";
            display: block;
            font-weight: 600;
            margin-bottom: 5px;
        }
    </style>
""", unsafe_allow_html=True)

# --- WEB UI ---
st.title("🇬🇹 MAGA: Procesador de Facturas por la LAE")
uploaded_pdfs = st.file_uploader(label='1. Seleccione sus Facturas (PDFs)', type='pdf', accept_multiple_files=True,
                                 help='Arrastre y suelte sus facturas aquí. El límite es 200mb por archivo')
uploaded_xlsx = st.file_uploader(label='2. Seleccione su Archivo de Excel', type='xlsx',
                                 help='Arrastre y suelte el archivo de Excel dónde van los totales de las facturas')

if st.button("INICIAR PROCESO") and uploaded_pdfs and uploaded_xlsx:
    try:
        input_buffer = io.BytesIO(uploaded_xlsx.read())
        wb = openpyxl.load_workbook(input_buffer)
        ws = wb.active 
        
        # 1. Setup "Extra Detalles" Sheet
        if "Extra Detalles" not in wb.sheetnames:
            ws_det = wb.create_sheet("Extra Detalles")
            ws_det.append(['Nombre Emisor', 'NIT Emisor', 'NIT Receptor', 'UUID', 'Municipio', 'Alerta % Abarrotes'])
        else:
            ws_det = wb["Extra Detalles"]

        # 2. Gather previously processed UUIDs (safely ignoring headers)
        processed_uuids = set()
        for row in ws_det.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
            if row[0] and str(row[0]).strip() != 'UUID': 
                processed_uuids.add(str(row[0]).strip())

        # 3. Map Excel Columns
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=15): 
            for cell in row:
                if not cell.value: continue
                val = normalize_text(str(cell.value))
                
                if 'abarrotes' in val: col_map['abar'] = cell.column
                if 'agricultura' in val: col_map['agri'] = cell.column
                if 'escuela' in val or 'establecimiento' in val: col_map['escuelas'] = cell.column
                
                # Logic for merged Proveedores header: finds the "Total" sub-column underneath
                if 'proveedor' in val or 'productor' in val:
                    base_col = cell.column
                    base_row = cell.row
                    found_total = False
                    
                    for r_offset in range(1, 4):
                        for c_offset in range(3):
                            sub_cell = ws.cell(row=base_row + r_offset, column=base_col + c_offset)
                            if sub_cell.value and 'total' in normalize_text(str(sub_cell.value)):
                                col_map['productores'] = sub_cell.column
                                found_total = True
                                break
                        if found_total: break
                    
                    if 'productores' not in col_map:
                        col_map['productores'] = base_col

        if 'abar' not in col_map or 'agri' not in col_map:
            st.error(f"No encontré las columnas base en el Excel. Columnas detectadas: {col_map}")
            st.stop()

        muni_map = {'totonicapan, totonicapan': 1, 'totonicapan': 1, 'san cristobal totonicapan': 2, 
                    'san francisco el alto': 3, 'san andres xecul': 4, 'momostenango': 5, 
                    'santa maria chiquimula': 6, 'santa lucia la reforma': 7, 'san bartolo': 8}
        
        # Sort municipalities by string length so it matches longest names (Santa Lucia) before short ones (Totonicapan)
        sorted_munis = sorted(muni_map.items(), key=lambda x: len(x[0]), reverse=True)

        batch_totals = {m_id: {'abar': 0.0, 'agri': 0.0, 'emisores': set(), 'receptores': set()} for m_id in muni_map.values()}
        new_count = 0
        skipped_count = 0
        progress_bar = st.progress(0)

        # 4. Process each PDF
        for i, pdf_file in enumerate(uploaded_pdfs):
            with pdfplumber.open(pdf_file) as pdf:
                text = "".join([p.extract_text() or "" for p in pdf.pages])
                tables = []
                for p in pdf.pages:
                    t = p.extract_table()
                    if t: tables.extend(t)

                # Strict Regex to ensure we don't accidentally grab a non-UUID number
                uuid_m = re.search(r'\b[A-F0-9]{8}-[A-F0-9]{4}-[A-F0-9]{4}-[A-F0-9]{4}-[A-F0-9]{12}\b', text, re.I)
                uuid_val = uuid_m.group(0).upper() if uuid_m else pdf_file.name

                if uuid_val in processed_uuids: 
                    st.info(f"⏭️ Factura omitida (ya sumada anteriormente): {pdf_file.name}")
                    skipped_count += 1
                    continue

                # Completely crush all text spaces, newlines, and commas for a 100% reliable municipality match
                text_squished = re.sub(r'[\s,]+', '', normalize_text(text))
                
                m_id = None
                m_name = "N/A"
                for k, v in sorted_munis:
                    key_squished = re.sub(r'[\s,]+', '', normalize_text(k))
                    if key_squished in text_squished:
                        m_id = v
                        m_name = k
                        break

                if m_id:
                    abar_sum, agri_sum = 0, 0
                    
                    # Expanded lists to ensure everything you process gets categorized
                    cultivados = ['tomate', 'pina', 'piña', 'banano', 'zanahoria', 'guisquil', 'cebolla', 'aguacate', 
                                  'miltomate', 'brocoli', 'melon', 'melón', 'ejote', 'maiz', 'maíz', 'jamaica', 
                                  'cebada', 'papaya', 'manzana', 'chile', 'apio', 'ajo', 'cilantro']
                    abarrotes = ['pollo', 'tostada', 'huevo', 'pan']
                    
                    total_col_idx = -1
                    for row_tbl in tables:
                        if not row_tbl: continue
                        for idx, cell in enumerate(row_tbl):
                            if cell and 'total' in normalize_text(str(cell)) and 'descuento' not in normalize_text(str(cell)):
                                total_col_idx = idx
                                break
                        if total_col_idx != -1: break

                    for row_tbl in tables:
                        if not row_tbl: continue
                        row_text = " ".join([normalize_text(str(x)) for x in row_tbl if x])
                        
                        val = 0.0
                        if total_col_idx != -1 and len(row_tbl) > total_col_idx:
                            val = clean_currency(row_tbl[total_col_idx])
                        else:
                            if len(row_tbl) >= 8: val = clean_currency(row_tbl[7])
                            elif len(row_tbl) >= 7: val = clean_currency(row_tbl[6])
                            elif len(row_tbl) >= 4: val = clean_currency(row_tbl[-1]) 
                            
                        if any(x in row_text for x in cultivados): agri_sum += val
                        if any(x in row_text for x in abarrotes): abar_sum += val
                    
                    nit_e_match = re.search(r'Emisor:\s*([0-9Kk\-]+)', text, re.I)
                    nit_r_match = re.search(r'Receptor:\s*([0-9Kk\-]+)', text, re.I)
                    
                    # DOTALL added here to fix issues where the name spills onto the next line
                    name_e_match = re.search(r'(?:Factura(?:\s*Pequeño\s*Contribuyente)?)\s*\n+(.*?)\n+Nit\s*Emisor', text, re.IGNORECASE | re.DOTALL)
                    
                    nit_e = nit_e_match.group(1).strip() if nit_e_match else "N/A"
                    nit_r = nit_r_match.group(1).strip() if nit_r_match else "N/A"
                    
                    raw_name = name_e_match.group(1).strip() if name_e_match else "N/A"
                    raw_name = re.sub(r'\s+', ' ', raw_name) # Clean up messy line breaks in name
                    name_e = re.split(r'(?i)n[úu]mero\s*de\s*autorizaci[óo]n', raw_name)[0]
                    name_e = re.split(r'(?i)\bserie\b', name_e)[0].strip()

                    batch_totals[m_id]['abar'] += abar_sum
                    batch_totals[m_id]['agri'] += agri_sum
                    if nit_e != "N/A": batch_totals[m_id]['emisores'].add(nit_e)
                    if nit_r != "N/A": batch_totals[m_id]['receptores'].add(nit_r)

                    total_rec = abar_sum + agri_sum
                    perc_abar = (abar_sum / total_rec) if total_rec > 0 else 0
                    alert_status = "⚠️ ALERTA: >30%" if perc_abar > 0.30 else "OK"

                    ws_det.append([name_e, nit_e, nit_r, uuid_val, m_name, alert_status])
                    processed_uuids.add(uuid_val)
                    new_count += 1
                else:
                    st.warning(f"No se pudo identificar el municipio en la factura: {pdf_file.name}")

            progress_bar.progress((i + 1) / len(uploaded_pdfs))

        # 5. Write accumulated data to Main Sheet safely
        for row_ex in ws.iter_rows(min_row=1, max_row=200):
            cell_a_val = str(row_ex[0].value).strip() if row_ex[0].value is not None else ""
            if not cell_a_val: continue

            try:
                excel_m_id = int(float(cell_a_val))
                if excel_m_id in batch_totals:
                    r_idx = row_ex[0].row
                    data = batch_totals[excel_m_id]

                    # Uses the new safe_float() to guarantee it can add to the existing cell format
                    if 'abar' in col_map and data['abar'] > 0:
                        curr_abar = ws.cell(r_idx, col_map['abar']).value
                        ws.cell(r_idx, col_map['abar']).value = safe_float(curr_abar) + data['abar']
                    
                    if 'agri' in col_map and data['agri'] > 0:
                        curr_agri = ws.cell(r_idx, col_map['agri']).value
                        ws.cell(r_idx, col_map['agri']).value = safe_float(curr_agri) + data['agri']

                    if 'escuelas' in col_map and len(data['receptores']) > 0:
                        curr_esc = ws.cell(r_idx, col_map['escuelas']).value
                        ws.cell(r_idx, col_map['escuelas']).value = int(safe_float(curr_esc)) + len(data['receptores'])
                    
                    if 'productores' in col_map and len(data['emisores']) > 0:
                        curr_prod = ws.cell(r_idx, col_map['productores']).value
                        ws.cell(r_idx, col_map['productores']).value = int(safe_float(curr_prod)) + len(data['emisores'])

            except (ValueError, TypeError):
                continue

        # 6. Format "Extra Detalles" (Auto-width and Borders)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col in ws_det.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column) 
            
            for cell in col:
                cell.border = thin_border 
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_det.column_dimensions[col_letter].width = max_length + 2

        # 7. Final Export
        output = io.BytesIO()
        wb.save(output)
        
        st.success(f"¡Proceso completado! {new_count} facturas nuevas procesadas y agregadas al Excel.")
        if skipped_count > 0:
            st.info(f"Nota: Se saltaron {skipped_count} facturas porque ya estaban registradas en el Excel.")
            
        output.seek(0)
        st.download_button("Descargar Reporte Final", data=output.getvalue(), 
                           file_name="Reporte_MAGA_Actualizado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"Error crítico detectado: {e}")
